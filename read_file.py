import openpyxl

# Load the workbook
workbook = openpyxl.load_workbook('./Pseudonymph-GENERIC-23012025-1439.xlsx')

# Function to print the content of a sheet
def print_sheet_content(sheet):
    for row in sheet.iter_rows(values_only=True):
        print(row)

# Print the content of each sheet
for sheet_name in workbook.sheetnames:
    print(f"Sheet: {sheet_name}")
    sheet = workbook[sheet_name]
    print_sheet_content(sheet)
    print("\n")